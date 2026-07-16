"""
Extractor de archivos Digiventure (formulario de solicitud en formato XLSX)
Lee los archivos DG_*.xlsx que contienen la información de la solicitud de crédito.

Estructura real del Excel Digiventure (verificada con archivos de muestra):
  - "Simulador": tabla de amortización pre-calculada + datos del préstamo
  - "Flujo Empleado": análisis financiero (ingresos, gastos, cuotas)
  - "Analisis": hoja de análisis con scoring, Begini, QUANTO, decisión
  - "Evidencia Fotografica": nombre, cédula, destino del crédito
  - "ListaDepegable": escala Begini de referencia
  - "Plan de pagos": tabla de pagos
  - "Guion" / "Preguntas_Reto" / "Referenciación": operativas
"""
import logging
from pathlib import Path
from typing import Optional

log = logging.getLogger(__name__)


def extraer_digiventure(filepath: str) -> dict:
    """
    Lee un archivo XLSX de Digiventure y extrae la información relevante.
    
    Args:
        filepath: Ruta al archivo .xlsx
    
    Returns:
        dict con datos de la solicitud, captura y análisis financiero
    """
    try:
        import openpyxl
    except ImportError:
        log.error("openpyxl no instalado. Ejecute: pip install openpyxl")
        return {}

    data = {
        # Datos de solicitud
        "cedula": None,
        "nombre_solicitante": None,
        "destino_credito": None,
        "monto_solicitado": None,
        "monto_aprobado": None,
        "neto_desembolsar": None,
        "linea_credito": None,
        "entidad_trabajo": None,
        "cargo": None,
        "analista": None,
        
        # Análisis / Scoring
        "score_begini": None,
        "score_centrales": None,
        "quanto_ingreso": None,
        "cuota_nueva_mas_gastos": None,
        "capacidad_pago": None,
        "decision_begini": None,
        "decision_centrales": None,
        "decision_quanto": None,
        "decision_cuota": None,
        "decision_capacidad": None,
        "estado_adres": None,
        "decision_adres": None,
        "concepto_ejecutivo": None,
        
        # Flujo financiero
        "ingreso_reportado": None,
        "quanto_ingreso_flujo": None,
        "ingresos_brutos": None,
        "comisiones": None,
        "arriendos": None,
        "otros_ingresos": None,
        "total_ingresos": None,
        "descuentos_ley": None,
        "ingresos_netos": None,
        "gastos_sostenimiento": None,
        "arrendamientos": None,
        "manutencion_hijos": None,
        "cuotas_creditos": None,
        "cuotas_tarjetas": None,
        "cuotas_deudas_particulares": None,
        "otros_gastos": None,
        "total_gastos": None,
        "disponible": None,
        
        # Simulador
        "valor_solicitado_sim": None,
        "interes_total": None,
        "seguro_vida_total": None,
        "fianza_iva": None,
        "tecnologia_iva": None,
        "administracion": None,
        "descuento_inclusion": None,
        "total_a_pagar": None,
    }

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        log.error(f"Error abriendo XLSX: {e}")
        return data

    # ── Helper para leer celdas de forma segura ─────────────────────────────
    def cell_val(ws, cell_ref, default=None):
        try:
            val = ws[cell_ref].value
            return val if val is not None else default
        except Exception:
            return default

    # ── Evidencia Fotografica ───────────────────────────────────────────────
    ev_names = ["Evidencia Fotografica", "Evidencia Fotográfica"]
    for ev_name in ev_names:
        if ev_name in wb.sheetnames:
            ws = wb[ev_name]
            data["nombre_solicitante"] = cell_val(ws, "C4")
            data["cedula"] = cell_val(ws, "C5")
            data["destino_credito"] = cell_val(ws, "C6")
            break

    # ── Analisis ────────────────────────────────────────────────────────────
    analisis_names = ["Analisis", "Análisis"]
    for an_name in analisis_names:
        if an_name in wb.sheetnames:
            ws = wb[an_name]
            
            data["analista"] = cell_val(ws, "C4")
            
            # Fallback: nombre y cédula
            if not data["nombre_solicitante"]:
                data["nombre_solicitante"] = cell_val(ws, "C6")
            if not data["cedula"]:
                data["cedula"] = cell_val(ws, "C7")
            
            data["entidad_trabajo"] = cell_val(ws, "E6")
            data["cargo"] = cell_val(ws, "E7")
            data["linea_credito"] = cell_val(ws, "E8")
            
            # Parámetros de aprobación/rechazo
            data["score_begini"] = cell_val(ws, "C12")
            data["decision_begini"] = cell_val(ws, "D12")
            data["score_centrales"] = cell_val(ws, "C13")
            data["decision_centrales"] = cell_val(ws, "D13")
            data["quanto_ingreso"] = cell_val(ws, "C14")
            data["decision_quanto"] = cell_val(ws, "D14")
            data["cuota_nueva_mas_gastos"] = cell_val(ws, "C15")
            data["decision_cuota"] = cell_val(ws, "D15")
            data["capacidad_pago"] = cell_val(ws, "C16")
            data["decision_capacidad"] = cell_val(ws, "D16")
            data["estado_adres"] = cell_val(ws, "C17")
            data["decision_adres"] = cell_val(ws, "D17")
            
            # Solicitud
            data["monto_solicitado"] = cell_val(ws, "C21")
            data["concepto_ejecutivo"] = cell_val(ws, "D21")
            data["monto_aprobado"] = cell_val(ws, "C22")
            data["neto_desembolsar"] = cell_val(ws, "C25")
            break

    # ── Flujo Empleado ──────────────────────────────────────────────────────
    flujo_names = ["Flujo Empleado", "Flujo credito", "Flujo Credito"]
    for fl_name in flujo_names:
        if fl_name in wb.sheetnames:
            ws = wb[fl_name]
            
            # Fallback: nombre y cédula
            if not data["nombre_solicitante"]:
                data["nombre_solicitante"] = cell_val(ws, "C4")
            if not data["cedula"]:
                data["cedula"] = cell_val(ws, "C5")
            
            data["ingreso_reportado"] = cell_val(ws, "D8")
            data["quanto_ingreso_flujo"] = cell_val(ws, "D9")
            data["ingresos_brutos"] = cell_val(ws, "D10")
            data["comisiones"] = cell_val(ws, "D11")
            data["arriendos"] = cell_val(ws, "D12")
            data["otros_ingresos"] = cell_val(ws, "D14")
            data["total_ingresos"] = cell_val(ws, "D15")
            data["descuentos_ley"] = cell_val(ws, "D16")
            data["ingresos_netos"] = cell_val(ws, "D17")
            
            data["gastos_sostenimiento"] = cell_val(ws, "D19")
            data["arrendamientos"] = cell_val(ws, "D20")
            data["cuotas_creditos"] = cell_val(ws, "D22")
            data["cuotas_tarjetas"] = cell_val(ws, "D23")
            data["cuotas_deudas_particulares"] = cell_val(ws, "D24")
            data["otros_gastos"] = cell_val(ws, "D25")
            
            # Buscar total gastos y disponible (pueden estar en filas variables)
            for row_num in range(26, 40):
                label = cell_val(ws, f"C{row_num}")
                val = cell_val(ws, f"D{row_num}")
                if label:
                    label_lower = str(label).lower().strip()
                    if "total" in label_lower and "gasto" in label_lower:
                        data["total_gastos"] = val
                    elif "disponible" in label_lower:
                        data["disponible"] = val
            break

    # ── Simulador ───────────────────────────────────────────────────────────
    if "Simulador" in wb.sheetnames:
        ws = wb["Simulador"]
        data["valor_solicitado_sim"] = cell_val(ws, "B15")
        data["interes_total"] = cell_val(ws, "B16")
        data["seguro_vida_total"] = cell_val(ws, "B17")
        data["fianza_iva"] = cell_val(ws, "B18")
        data["tecnologia_iva"] = cell_val(ws, "B19")
        data["administracion"] = cell_val(ws, "B20")
        data["descuento_inclusion"] = cell_val(ws, "B21")
        data["total_a_pagar"] = cell_val(ws, "B22")

    # ── Convertir cédula a string ───────────────────────────────────────────
    if data["cedula"]:
        data["cedula"] = str(int(data["cedula"])) if isinstance(data["cedula"], (int, float)) else str(data["cedula"]).strip()

    # ── Limpiar nombre ──────────────────────────────────────────────────────
    if data["nombre_solicitante"]:
        data["nombre_solicitante"] = str(data["nombre_solicitante"]).strip()

    wb.close()

    log.info(
        f"Digiventure extraido: CC={data['cedula']} Nombre={data['nombre_solicitante']} "
        f"Monto={data['monto_solicitado']} Aprobado={data['monto_aprobado']} "
        f"Score={data['score_centrales']} Begini={data['score_begini']}"
    )
    return data
